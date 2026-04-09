"""
seed_usuarios_prod.py — Carga 1,817 tiendas reales + usuarios desde Tienda.xlsx
================================================================================
Uso:
  python scripts/seed_usuarios_prod.py --excel ruta/Tienda.xlsx [--confirm]

Es idempotente: usa ON CONFLICT para upsert. Commit por fila para evitar
que un error aborte toda la transacción.
"""

import sys, os, argparse

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from sqlalchemy import text
from app.db.session import SessionLocal, engine
from app.models.models import Base, Tienda, Usuario, RolUsuario, Zona, Region


def get_or_create_region(db, nombre):
    r = db.query(Region).filter(Region.nombre == nombre).first()
    if not r:
        r = Region(nombre=nombre)
        db.add(r)
        db.commit()
        db.refresh(r)
    return r


def get_or_create_zona(db, nombre, region_id):
    z = (
        db.query(Zona)
        .filter(Zona.nombre == nombre, Zona.region_id == region_id)
        .first()
    )
    if not z:
        z = Zona(nombre=nombre, region_id=region_id)
        db.add(z)
        db.commit()
        db.refresh(z)
    return z


def upsert_tienda(db, economico, nombre, zona_id, correo):
    db.execute(
        text(
            """
        INSERT INTO tiendas (nombre, zona_id, correo_corporativo, centro_costos, activo)
        VALUES (:nombre, :zona_id, :correo, :eco, true)
        ON CONFLICT (correo_corporativo) DO UPDATE
          SET nombre        = EXCLUDED.nombre,
              zona_id       = EXCLUDED.zona_id,
              centro_costos = EXCLUDED.centro_costos
    """
        ),
        {"nombre": nombre, "zona_id": zona_id, "correo": correo, "eco": economico},
    )
    db.commit()
    return db.query(Tienda).filter(Tienda.correo_corporativo == correo).first()


def upsert_usuario(db, email, nombre, hashed_pw, tienda_id):
    db.execute(
        text(
            """
        INSERT INTO usuarios (email, nombre, hashed_password, rol, tienda_id, activo)
        VALUES (:email, :nombre, :pw, 'TIENDA', :tid, true)
        ON CONFLICT (email) DO UPDATE
          SET nombre          = EXCLUDED.nombre,
              hashed_password = EXCLUDED.hashed_password,
              tienda_id       = EXCLUDED.tienda_id
    """
        ),
        {"email": email, "nombre": nombre, "pw": hashed_pw, "tid": tienda_id},
    )
    db.commit()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="scripts/Tienda.xlsx")
    parser.add_argument("--confirm", action="store_true")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"ERROR: No se encontró: {args.excel}")
        sys.exit(1)

    print(f"Leyendo {args.excel}...")
    wb = load_workbook(args.excel, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    IDX = {h: i for i, h in enumerate(header)}

    activas = [r for r in rows[1:] if r[IDX.get("status", 4)] == "ACTIVA"]
    print(
        f"  ACTIVAS: {len(activas)}  |  SINIESTRADAS omitidas: {len(rows)-1-len(activas)}"
    )

    if not args.confirm:
        print("\nPREVIEW — agrega --confirm para aplicar")
        print(f"  Se procesarían {len(activas)} tiendas y {len(activas)} usuarios")
        print("\nMuestra primeras 5:")
        for r in activas[:5]:
            print(
                f"  {str(r[IDX.get('economico',1)]):6s}  {str(r[IDX.get('correo_corpo',13)]):35s}  pass: {r[IDX.get('password_plain',16)]}"
            )
        return

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    region_cache = {}
    zona_cache = {}
    ok = err = 0

    print(f"\nProcesando {len(activas)} tiendas...")

    for i, r in enumerate(activas, 1):
        try:
            economico = str(r[IDX.get("economico", 1)] or "").strip()
            nombre_neto = str(r[IDX.get("nombre_neto", 2)] or "").strip()
            nombre_region = str(r[IDX.get("nombre_region", 9)] or "SIN REGION").strip()
            correo = str(r[IDX.get("correo_corpo", 13)] or "").strip().lower()
            hashed_pw = str(r[IDX.get("password", 15)] or "").strip()

            if not correo or not economico:
                err += 1
                continue

            if nombre_region not in region_cache:
                region_cache[nombre_region] = get_or_create_region(db, nombre_region)

            zona_key = nombre_region
            if zona_key not in zona_cache:
                zona_cache[zona_key] = get_or_create_zona(
                    db, nombre_region, region_cache[nombre_region].id
                )

            tienda = upsert_tienda(
                db, economico, nombre_neto, zona_cache[zona_key].id, correo
            )
            upsert_usuario(db, correo, nombre_neto, hashed_pw, tienda.id)

            ok += 1
            if i % 100 == 0:
                print(f"  {i}/{len(activas)} procesadas  ({err} errores)")

        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            print(f"  ERROR fila {i} eco={r[1]}: {e}")
            err += 1

    db.close()

    print(
        f"""
{'='*50}
COMPLETADO
{'='*50}
  Procesadas OK : {ok}
  Errores       : {err}

Credenciales:
  Email   : {{economico}}@soyneto.com  (ej: 100@soyneto.com)
  Password: N3T0{{economico}}           (ej: N3T0100)
"""
    )


if __name__ == "__main__":
    main()
